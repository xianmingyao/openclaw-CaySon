# -*- coding: utf-8 -*-
import os
import openpyxl

# 列出jingmai-product-publish目录下的所有文件
folder = r'E:\workspace\skills\jingmai-product-publish'
print("Files in jingmai-product-publish folder:")
for f in os.listdir(folder):
    full_path = os.path.join(folder, f)
    print(f"  {f} ({os.path.getsize(full_path)} bytes)")

# 找xlsx文件
xlsx_files = [f for f in os.listdir(folder) if f.endswith('.xlsx')]
print(f"\nExcel files: {xlsx_files}")

# 读取Excel
if xlsx_files:
    for xlsx_file in xlsx_files:
        xlsx_path = os.path.join(folder, xlsx_file)
        print(f"\n{'='*60}")
        print(f"Reading: {xlsx_file}")
        print('='*60)
        
        wb = openpyxl.load_workbook(xlsx_path)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\nSheet: {sheet_name}")
            
            # 读取前5行
            for row_idx, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
                row_values = [str(v)[:30] if v else '' for v in row[:12]]
                print(f"  Row {row_idx}: {row_values}")
            
            # 找京东链接
            print(f"\n  京东链接商品:")
            for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), 4):
                if row[11] and 'jd.com' in str(row[11]):
                    print(f"    Row {row_idx}:")
                    print(f"      商品名称: {row[2]}")
                    print(f"      品牌: {row[3]}")
                    print(f"      型号: {row[4]}")
                    print(f"      链接: {row[11]}")
