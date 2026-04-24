# -*- coding: utf-8 -*-
import openpyxl
import os

xlsx_path = r'E:\workspace\skills\jingmai-product-publish\湖南上架表格.xlsx'

print(f"Reading: {xlsx_path}")

# 加载工作簿
wb = openpyxl.load_workbook(xlsx_path)

print(f"\nSheet names: {wb.sheetnames}")

# 遍历所有sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*60}")
    
    # 读取前20行
    for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
        if any(cell is not None for cell in row):
            print(f"Row {i+1}: {row}")
