# -*- coding: utf-8 -*-
import openpyxl
import sys
import io

# 设置stdout编码
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

xlsx_path = r'E:\workspace\skills\jingmai-product-publish\湖南上架表格.xlsx'

print(f"Reading: {xlsx_path}")

wb = openpyxl.load_workbook(xlsx_path)

print(f"\nSheet names:")
for name in wb.sheetnames:
    print(f"  - {name}")

# 读取上架模板sheet
ws = wb['上架模板']
print(f"\n{'='*80}")
print("上架模板 Sheet")
print('='*80)

# 读取表头和数据
headers = []
for row_idx, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
    row_data = []
    for cell in row:
        if cell is not None:
            row_data.append(str(cell))
        else:
            row_data.append('')
    print(f"Row {row_idx}: {row_data}")

# 找到京东链接列
print("\n" + "="*80)
print("商品链接信息:")
print("="*80)

# 遍历所有行，查找京东链接
for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), 4):
    # row[11] 是第12列，京东链接
    if row[11]:  # 京东链接列
        link = row[11]
        if isinstance(link, str) and 'jd.com' in link:
            print(f"\nRow {row_idx}:")
            print(f"  商品名称: {row[2]}")  # 商品名称
            print(f"  品牌: {row[3]}")     # 品牌
            print(f"  型号: {row[4]}")     # 商品型号
            print(f"  京东链接: {link}")    # 京东链接
            
            # 提取商品ID
            import re
            match = re.search(r'/(\d+)\.html', link)
            if match:
                print(f"  商品ID: {match.group(1)}")
